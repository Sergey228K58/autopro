from flask import Flask, render_template, request, redirect, url_for, session, send_file
from openpyxl import load_workbook, Workbook
from pathlib import Path
from functools import wraps
import bcrypt
import io
import datetime

app = Flask(__name__)
app.secret_key = "change_this_secret_key"

DATA_DIR = Path("data")
DATA_DIR.mkdir(exist_ok=True)

USERS_FILE = DATA_DIR / "users.xlsx"
USER_HEADERS = ["id", "username", "password_hash", "role"]

ORDERS_FILE = DATA_DIR / "orders.xlsx"
ORDER_HEADERS = [
    "id", "client_name", "client_username",
    "car_brand", "car_model", "year", "config",
    "color", "price", "status"
]

SERVICE_FILE = DATA_DIR / "service_orders.xlsx"
SERVICE_HEADERS = [
    "id", "client_username", "vin", "works",
    "hours", "work_cost", "parts_cost",
    "total_cost", "status", "master", "finished_at"
]

PARTS_FILE = DATA_DIR / "parts.xlsx"
PARTS_HEADERS = [
    "id", "name", "article", "category",
    "supplier", "qty", "min_qty", "price"
]

FINANCE_FILE = DATA_DIR / "finance.xlsx"
FINANCE_HEADERS = [
    "id", "type", "category", "amount",
    "related_order_id", "created_at"
]

CHATS_FILE = DATA_DIR / "chats.xlsx"
CHAT_HEADERS = ["id", "client_username", "manager_username", "message", "sender", "created_at"]

BOOKINGS_FILE = DATA_DIR / "bookings.xlsx"
BOOKING_HEADERS = [
    "id", "client_username", "date", "time",
    "service_type", "status"
]


def get_or_create_wb(path, headers):
    if not path.exists():
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        wb.save(path)
    return load_workbook(path)


def get_or_create_users():
    if not USERS_FILE.exists():
        wb = Workbook()
        ws = wb.active
        ws.append(USER_HEADERS)
        password_hash = bcrypt.hashpw("admin".encode(), bcrypt.gensalt()).decode()
        ws.append([1, "admin", password_hash, "admin"])
        wb.save(USERS_FILE)
    return load_workbook(USERS_FILE)


def find_user(username):
    wb = get_or_create_users()
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == username:
            return {
                "id": row[0],
                "username": row[1],
                "password_hash": row[2],
                "role": row[3]
            }
    return None


def role_required(*roles):
    def wrapper(func):
        @wraps(func)
        def decorated_view(*args, **kwargs):
            user = session.get("user")
            if not user:
                return redirect(url_for("login"))
            if roles and user.get("role") not in roles:
                return "Доступ запрещён"
            return func(*args, **kwargs)
        return decorated_view
    return wrapper


@app.route("/")
@role_required()
def index():
    return render_template("index.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"].encode()
        user = find_user(username)
        if user and bcrypt.checkpw(password, user["password_hash"].encode()):
            session["user"] = {
                "username": user["username"],
                "role": user["role"]
            }
            return redirect(url_for("index"))
        return render_template("login.html", error="Неверный логин или пароль")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/orders")
@role_required("manager", "admin")
def orders_list():
    wb = get_or_create_wb(ORDERS_FILE, ORDER_HEADERS)
    ws = wb.active
    orders = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        orders.append({
            "id": row[0],
            "client_name": row[1],
            "client_username": row[2],
            "car_brand": row[3],
            "car_model": row[4],
            "year": row[5],
            "config": row[6],
            "color": row[7],
            "price": row[8],
            "status": row[9],
        })
    statuses = [
        "Согласован",
        "Заказан у поставщика",
        "В пути",
        "На СТО",
        "ТО пройден",
        "Готов к выдаче"
    ]
    return render_template("orders_list.html", orders=orders, statuses=statuses)


@app.route("/orders/create", methods=["GET", "POST"])
@role_required("manager", "admin")
def order_create():
    if request.method == "POST":
        client_name = request.form["client_name"]
        client_username = request.form["client_username"]
        car_brand = request.form["car_brand"]
        car_model = request.form["car_model"]
        year = request.form["year"]
        config = request.form["config"]
        color = request.form["color"]
        price = request.form["price"]

        wb = get_or_create_wb(ORDERS_FILE, ORDER_HEADERS)
        ws = wb.active
        new_id = ws.max_row
        status = "Согласован"
        ws.append([
            new_id, client_name, client_username,
            car_brand, car_model, year, config,
            color, price, status
        ])
        wb.save(ORDERS_FILE)
        return redirect(url_for("orders_list"))
    return render_template("order_create.html")


@app.route("/orders/<int:order_id>/status", methods=["POST"])
@role_required("manager", "admin")
def order_update_status(order_id):
    new_status = request.form["status"]
    wb = get_or_create_wb(ORDERS_FILE, ORDER_HEADERS)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == order_id:
            row[9].value = new_status
            break
    wb.save(ORDERS_FILE)
    return redirect(url_for("orders_list"))


@app.route("/orders/<int:order_id>/contract")
@role_required("manager", "admin")
def order_contract(order_id):
    wb = get_or_create_wb(ORDERS_FILE, ORDER_HEADERS)
    ws = wb.active
    order = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == order_id:
            order = row
            break
    if not order:
        return "Заказ не найден"
    text = f"""ДОГОВОР КУПЛИ-ПРОДАЖИ АВТОМОБИЛЯ №{order[0]}

Покупатель: {order[1]} ({order[2]})
Автомобиль: {order[3]} {order[4]}, {order[5]} г.в., {order[6]}, цвет {order[7]}
Цена: {order[8]} руб.

Стороны договорились о передаче автомобиля после прохождения всех этапов подготовки.
"""
    return send_file(
        io.BytesIO(text.encode("utf-8")),
        mimetype="text/plain",
        as_attachment=True,
        download_name=f"contract_{order_id}.txt"
    )



@app.route("/service")
@role_required("master", "manager", "admin")
def service_list():
    wb = get_or_create_wb(SERVICE_FILE, SERVICE_HEADERS)
    ws = wb.active
    orders = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        orders.append({
            "id": row[0],
            "client_username": row[1],
            "vin": row[2],
            "works": row[3],
            "hours": row[4],
            "work_cost": row[5],
            "parts_cost": row[6],
            "total_cost": row[7],
            "status": row[8],
            "master": row[9],
            "finished_at": row[10],
        })
    return render_template("service_list.html", service_orders=orders)


@app.route("/service/create", methods=["GET", "POST"])
@role_required("master", "manager", "admin")
def service_create():
    if request.method == "POST":
        client_username = request.form["client_username"]
        vin = request.form["vin"]
        works = request.form["works"]
        hours = float(request.form["hours"])
        work_cost = float(request.form["work_cost"])
        parts_cost = float(request.form["parts_cost"])
        total_cost = work_cost + parts_cost
        status = "Открыт"
        master = session["user"]["username"]

        wb = get_or_create_wb(SERVICE_FILE, SERVICE_HEADERS)
        ws = wb.active
        new_id = ws.max_row
        ws.append([
            new_id, client_username, vin, works,
            hours, work_cost, parts_cost,
            total_cost, status, master, ""
        ])
        wb.save(SERVICE_FILE)
        return redirect(url_for("service_list"))
    return render_template("service_create.html")


@app.route("/service/<int:service_id>/close", methods=["POST"])
@role_required("master", "manager", "admin")
def service_close(service_id):
    wb = get_or_create_wb(SERVICE_FILE, SERVICE_HEADERS)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == service_id:
            row[8].value = "Завершён"
            row[10].value = datetime.datetime.now().isoformat(sep=" ", timespec="seconds")
            break
    wb.save(SERVICE_FILE)
    return redirect(url_for("service_list"))



@app.route("/parts")
@role_required("storekeeper", "master", "manager", "admin")
def parts_list():
    wb = get_or_create_wb(PARTS_FILE, PARTS_HEADERS)
    ws = wb.active
    parts = []
    low = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        item = {
            "id": row[0],
            "name": row[1],
            "article": row[2],
            "category": row[3],
            "supplier": row[4],
            "qty": row[5],
            "min_qty": row[6],
            "price": row[7],
        }
        parts.append(item)
        if item["qty"] is not None and item["min_qty"] is not None and item["qty"] <= item["min_qty"]:
            low.append(item)
    return render_template("parts_list.html", parts=parts, low_parts=low)


@app.route("/parts/receive", methods=["GET", "POST"])
@role_required("storekeeper", "admin")
def parts_receive():
    if request.method == "POST":
        name = request.form["name"]
        article = request.form["article"]
        category = request.form["category"]
        supplier = request.form["supplier"]
        qty = int(request.form["qty"])
        min_qty = int(request.form["min_qty"])
        price = float(request.form["price"])

        wb = get_or_create_wb(PARTS_FILE, PARTS_HEADERS)
        ws = wb.active

        # поиск по артикулу
        found = False
        for row in ws.iter_rows(min_row=2):
            if row[2].value == article:
                row[5].value = (row[5].value or 0) + qty
                row[7].value = price
                found = True
                break
        if not found:
            new_id = ws.max_row
            ws.append([new_id, name, article, category, supplier, qty, min_qty, price])

        wb.save(PARTS_FILE)
        return redirect(url_for("parts_list"))
    return render_template("parts_receive.html")


@app.route("/parts/inventory", methods=["POST"])
@role_required("storekeeper", "admin")
def parts_inventory():
    # для простоты — ничего не меняем, только "акт" в виде Excel
    wb = get_or_create_wb(PARTS_FILE, PARTS_HEADERS)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="inventory.xlsx"
    )



@app.route("/finance")
@role_required("accountant", "admin")
def finance_list():
    wb = get_or_create_wb(FINANCE_FILE, FINANCE_HEADERS)
    ws = wb.active
    ops = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        ops.append({
            "id": row[0],
            "type": row[1],
            "category": row[2],
            "amount": row[3],
            "related_order_id": row[4],
            "created_at": row[5],
        })
    return render_template("finance_list.html", operations=ops)


@app.route("/finance/add", methods=["POST"])
@role_required("accountant", "admin")
def finance_add():
    type_ = request.form["type"]
    category = request.form["category"]
    amount = float(request.form["amount"])
    related_order_id = request.form.get("related_order_id") or ""
    created_at = datetime.datetime.now().isoformat(sep=" ", timespec="seconds")

    wb = get_or_create_wb(FINANCE_FILE, FINANCE_HEADERS)
    ws = wb.active
    new_id = ws.max_row
    ws.append([new_id, type_, category, amount, related_order_id, created_at])
    wb.save(FINANCE_FILE)
    return redirect(url_for("finance_list"))


@app.route("/finance/calc_cost/<int:order_id>")
@role_required("accountant", "admin")
def finance_calc_cost(order_id):
    # для примера: ищем заказ и сервис, считаем сумму
    wb_orders = get_or_create_wb(ORDERS_FILE, ORDER_HEADERS)
    ws_o = wb_orders.active
    price = 0
    for row in ws_o.iter_rows(min_row=2, values_only=True):
        if row[0] == order_id:
            price = float(row[8] or 0)
            break

    wb_service = get_or_create_wb(SERVICE_FILE, SERVICE_HEADERS)
    ws_s = wb_service.active
    service_total = 0
    for row in ws_s.iter_rows(min_row=2, values_only=True):
        if str(row[0]) == str(order_id):
            service_total += float(row[7] or 0)

    total_cost = price + service_total
    return f"Себестоимость заказа {order_id}: {total_cost} руб."


@app.route("/finance/report")
@role_required("accountant", "admin")
def finance_report():
    wb = get_or_create_wb(FINANCE_FILE, FINANCE_HEADERS)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="finance_report.xlsx"
    )



@app.route("/client-portal")
@role_required("client", "admin")
def client_portal():
    user = session.get("user")
    username = user["username"]

    # заказы клиента
    wb_o = get_or_create_wb(ORDERS_FILE, ORDER_HEADERS)
    ws_o = wb_o.active
    orders = []
    for row in ws_o.iter_rows(min_row=2, values_only=True):
        if row[2] == username:
            orders.append(row)

    # сервисные заказы клиента
    wb_s = get_or_create_wb(SERVICE_FILE, SERVICE_HEADERS)
    ws_s = wb_s.active
    service_orders = []
    for row in ws_s.iter_rows(min_row=2, values_only=True):
        if row[1] == username:
            service_orders.append(row)

    # записи на ремонт
    wb_b = get_or_create_wb(BOOKINGS_FILE, BOOKING_HEADERS)
    ws_b = wb_b.active
    bookings = []
    for row in ws_b.iter_rows(min_row=2, values_only=True):
        if row[1] == username:
            bookings.append(row)

    return render_template(
        "client_portal.html",
        user=user,
        orders=orders,
        service_orders=service_orders,
        bookings=bookings
    )


@app.route("/client-portal/booking", methods=["GET", "POST"])
@role_required("client", "admin")
def booking():
    if request.method == "POST":
        user = session.get("user")
        username = user["username"]
        date = request.form["date"]
        time = request.form["time"]
        service_type = request.form["service_type"]
        status = "Новая"

        wb = get_or_create_wb(BOOKINGS_FILE, BOOKING_HEADERS)
        ws = wb.active
        new_id = ws.max_row
        ws.append([new_id, username, date, time, service_type, status])
        wb.save(BOOKINGS_FILE)
        return redirect(url_for("client_portal"))
    return render_template("booking.html")


@app.route("/chat", methods=["GET", "POST"])
@role_required("client", "manager", "admin")
def chat():
    user = session.get("user")
    username = user["username"]
    role = user["role"]

    wb = get_or_create_wb(CHATS_FILE, CHAT_HEADERS)
    ws = wb.active

    if request.method == "POST":
        message = request.form["message"]
        sender = role
        manager_username = ""
        client_username = ""
        if role == "client":
            client_username = username
        else:
            manager_username = username
        new_id = ws.max_row
        created_at = datetime.datetime.now().isoformat(sep=" ", timespec="seconds")
        ws.append([new_id, client_username, manager_username, message, sender, created_at])
        wb.save(CHATS_FILE)
        return redirect(url_for("chat"))

    messages = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        messages.append({
            "id": row[0],
            "client_username": row[1],
            "manager_username": row[2],
            "message": row[3],
            "sender": row[4],
            "created_at": row[5],
        })
    return render_template("chat.html", messages=messages)


if __name__ == "__main__":
    app.run(debug=True)
